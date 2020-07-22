from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
import datetime as dy
import requests
import urllib.request
import os 

sd_data = ''
get_sd_url = f'http://jeil.jje.hs.kr/jeil-h/0208/board/16996/'

ua = UserAgent()
header = {'User-Agent':str(ua.chrome)}
req_html = requests.get(get_sd_url, headers=header)
html = req_html.text

soup = BeautifulSoup(html,'html.parser')

for i in soup.find('tbody').find_all('a'):
    if f'{dy.datetime.today().month}/{dy.datetime.today().day}' in str(i):
        sd_data = i

data_url = str(sd_data.get('onclick'))[str(sd_data.get('onclick')).index('(')+1:str(sd_data.get('onclick')).index(')')].split(',')[1].replace("'",'')


url_sd = f'http://jeil.jje.hs.kr/jeil-h/0208/board/16996/{data_url}'
print(url_sd)


req_html = requests.get(url_sd, headers=header)
html = req_html.text

soup = BeautifulSoup(html,'html.parser')
pr = soup.find('dd').find_all('a')[1].get('href')

sh_name_day = soup.find('dd').find_all('a')[0].get_text()[soup.find('dd').find_all('a')[0].get_text().index("(") : soup.find('dd').find_all('a')[0].get_text().index(")")+1]
sh_name = f'{dy.datetime.today().month}.{dy.datetime.today().day}{sh_name_day}'
# sh_name = '날짜'
# print(sh_name)


preview_url = f'http://jeil.jje.hs.kr{pr}'

urllib.request.urlretrieve(preview_url, 'sd.xlsx')

data = load_workbook("sd.xlsx", data_only=True)
ds = data.active

wb = Workbook()
ws = wb.active

sum_cell = list()


for i in range(1,3):
    for j in range(3,11):
        sst = ''
        ws.cell(row=j-1, column=i).value = ds.cell(row=j, column=i).value 
        ws.cell(row=j-1, column=i).font = Font(name='맑은 고딕', size=13, bold=True)
        ws.cell(row=j-1, column=i).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        ws.column_dimensions['B'].width = 20

for i in range(15,28):
    for j in range(3,11):
        if j != 10:
            ws.row_dimensions[j].height = 36
        ws.cell(row=j-1, column=i-12).value = str(ds.cell(row=j, column=i).value)[0:2] + str(ds.cell(row=j, column=i).value)[2:]
        ws.cell(row=j-1, column=i-12).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
        ws.cell(row=j-1, column=i-12).font = Font(name='맑은 고딕', size=10, bold=True)
        if ds.cell(row=j, column=i).value == None:
            sum_cell.append([i-12, j])

# print(sum_cell)

if sum_cell != []:
    ws.merge_cells(start_row= sum_cell[0][1]-1, start_column=sum_cell[0][0]-1,end_row= sum_cell[0][1]-1,end_column=sum_cell[len(sum_cell)-1][0])

wb.save(f"시간표 {sh_name}.xlsx")

os.remove('sd.xlsx')
